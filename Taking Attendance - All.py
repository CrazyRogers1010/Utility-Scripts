
import os
import time
import datetime
import openpyxl
import datetime
import shutil

def list_to_string(list):
    string = "\""
    for item in list:
        string += item
        string += ", "
    string += "\""
    return string
def select_course_attendance_sheet(course):
    ap1_daily_attendance_question_spreadsheet_path = "/Users/christopherrogers/OneDrive - Collierville Schools/20-21/Classes/Attendance/AP1 - Daily Attendance Question.xlsx"
    phy_daily_attendance_question_spreadsheet_path = "/Users/christopherrogers/OneDrive - Collierville Schools/20-21/Classes/Attendance/PHY - Daily Attendance Question.xlsx"
    apc_daily_attendance_question_spreadsheet_path = "/Users/christopherrogers/OneDrive - Collierville Schools/20-21/Classes/Attendance/APC - Daily Attendance Question.xlsx"

    daily_attendance_question_spreadsheet_path = ""

    if course == "ap1":
        daily_attendance_question_spreadsheet_path = ap1_daily_attendance_question_spreadsheet_path
    if course == "phy":
        daily_attendance_question_spreadsheet_path = phy_daily_attendance_question_spreadsheet_path
    if course == "apc":
        daily_attendance_question_spreadsheet_path = apc_daily_attendance_question_spreadsheet_path

    return daily_attendance_question_spreadsheet_path

now = datetime.datetime.now()
print(now)


print(now.date())


dict_of_courses_and_dropped_students = {
    "ap1":["Firstname Lastname1","Firstname Lastname2","Firstname Lastname3"],
    "phy":["Firstname Lastname4","Firstname Lastname5","Firstname Lastname6"],
    "apc":[]
}

dict_of_courses_and_students_on_roster = {
    "1st": ["This contained a list of first and last names which was the roster for this class period"],
    "2nd": ["This contained a list of first and last names which was the roster for this class period"],
    "3rd": ["This contained a list of first and last names which was the roster for this class period"],
    "5th": ["This contained a list of first and last names which was the roster for this class period"],
    "7th": ["This contained a list of first and last names which was the roster for this class period"],
    "8th": ["This contained a list of first and last names which was the roster for this class period"]
}

dict_of_periods_and_courses = {
    "1st": "ap1",
    "2nd": "phy",
    "3rd": "apc",
    "5th": "ap1",
    "7th": "phy",
    "8th": "ap1"
}

period = ""
while period not in dict_of_courses_and_students_on_roster.keys():
    period = input("Which period? ")

course = ""

course = dict_of_periods_and_courses[period]
attendance_workbook_path = select_course_attendance_sheet(course)
attendance_workbook = openpyxl.load_workbook(attendance_workbook_path)


for sheet in attendance_workbook:
    if sheet.title == 'Form1':
        max_row = sheet.max_row
        rows_for_attendance = []
        for i in range(2,max_row + 1):
            start_time = sheet.cell(row = i, column = 2).value
            if start_time.date() == now.date():
                rows_for_attendance.append(i)
        list_of_present_students = []
        for row_num in rows_for_attendance:
            name = sheet.cell(row = row_num, column = 5).value
            if name not in list_of_present_students and name in dict_of_courses_and_students_on_roster[period]:
                list_of_present_students.append(name)
        print(len(list_of_present_students))

        list_of_all_names = []
        for i in range(2,max_row + 1):
            name = sheet.cell(row = i, column = 5).value
            if name not in list_of_all_names and name in dict_of_courses_and_students_on_roster[period] and name not in dict_of_courses_and_dropped_students[course]:
                list_of_all_names.append(name)
        print(len(list_of_all_names))
        print(len(dict_of_courses_and_students_on_roster[period]))



        list_absent_students = []
        for name in list_of_all_names:
            if name not in list_of_present_students:
                list_absent_students.append(name)

list_absent_students.sort()
for item in list_absent_students:
    print(item)


list_of_all_names.sort()
print(list_of_all_names)

