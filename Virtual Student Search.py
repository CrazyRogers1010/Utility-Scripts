
import os
import time
import openpyxl
import datetime
import shutil


roster_workbook = openpyxl.load_workbook('scoresheetReport (4).xlsx')

names_on_my_rosters = []
list_of_all_names = []
last_names = []
first_names = []
for sheet in roster_workbook.worksheets:
    list_of_students = []   
    max_row = sheet.max_row
    max_col = sheet.max_column
    for i in range(2,max_row + 1):
        names_on_my_rosters.append(sheet.cell(row = i, column = 1).value)
        list_of_students.append(sheet.cell(row = i, column = 1).value)
        list_of_all_names.append(sheet.cell(row = i, column = 1).value)
    for student_name in list_of_students:
        student_name_parts = student_name.split(',')
        last_names.append(student_name_parts[0])
        first_names.append(student_name_parts[1].strip())
print(len(last_names))
print(len(first_names))

print(len(list_of_all_names))


wb = openpyxl.load_workbook('/Users/christopherrogers/Desktop/Work Programs/CVA students (4).xlsx')
ws = wb.active

max_row = ws.max_row
list_of_virtual = []
counter = 0

print(ws.max_row)

for i in range(2,max_row+1):
    if ws.cell(row=i, column=1).value == None or ws.cell(row=i, column=2).value == None:
        print(i)
        continue
    else:
        if (ws.cell(row=i, column=1).value.strip() + ', ' + ws.cell(row=i, column=2).value.strip()) in list_of_all_names:
            list_of_virtual.append(ws.cell(row=i, column=1).value.strip() + ', ' + ws.cell(row=i, column=2).value.strip())
            counter += 1
print(counter)

virtual = []
in_person_students = []
for name in list_of_virtual:
    for sheet in roster_workbook.worksheets:
        list_of_students = []
        max_row = sheet.max_row
        max_col = sheet.max_column
        for i in range(2,max_row + 1):
            list_of_students.append(sheet.cell(row = i, column = 1).value)
        if name in list_of_students:
            virtual.append(sheet.title + ' - ' + name)
    if name not in list_of_students:
        in_person_students.append(sheet.title + ' - ' + name)
virtual.sort()
print('dfas ')
list_of_unneeded_periods = ['6(A) Study Skills A (Semester 1', '6(A) Study Skills A (Semester 2','HR(A) Homeroom 11']

for sheet in roster_workbook:
    if sheet.title not in list_of_unneeded_periods:
        list_of_students = []   
        max_row = sheet.max_row
        max_col = sheet.max_column
        for i in range(2,max_row + 1):
            list_of_students.append(sheet.cell(row = i, column = 1).value)
        for name in list_of_students:
            if name not in list_of_virtual:
                print(sheet.title + ' - ' + name)



    
